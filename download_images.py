#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
根据Excel数据批量下载图片，并按如下规则重命名保存：
文件名格式：一级-二级1-二级2-三级-品牌-条码.jpg
注意：下载前会去掉图片地址中的查询参数（?后面的内容）。
"""

import argparse
import os
import re
import sys
import threading
from typing import Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

LOG_DIR = os.path.join(os.getcwd(), "logs")
LOG_FILE = os.path.join(LOG_DIR, time.strftime("download-%Y%m%d.log"))
LOG_LOCK = threading.Lock()
ERROR_DIR = os.path.join(os.getcwd(), "error")
ERROR_FILE = os.path.join(ERROR_DIR, time.strftime("error-%Y%m%d.log"))
ERROR_LOCK = threading.Lock()

def write_log(text: str) -> None:
    try:
        os.makedirs(LOG_DIR, exist_ok=True)
        with LOG_LOCK:
            with open(LOG_FILE, "a", encoding="utf-8") as f:
                f.write(text + "\n")
    except Exception:
        pass

def write_error(text: str) -> None:
    try:
        os.makedirs(ERROR_DIR, exist_ok=True)
        with ERROR_LOCK:
            with open(ERROR_FILE, "a", encoding="utf-8") as f:
                f.write(text + "\n")
    except Exception:
        pass

import requests
from openpyxl import load_workbook


def sanitize(text: Optional[str]) -> str:
    """对文件名中的每个字段进行清洗，去除无效字符并压缩分隔符。"""
    if text is None:
        return ""
    s = str(text).strip()
    # 替换不允许的文件名字符为破折号
    s = re.sub(r"[\\/\\:*?\"<>|]", "-", s)
    # 去除多余空格
    s = re.sub(r"\s+", "", s)
    # 压缩多个破折号
    s = re.sub(r"-+", "-", s)
    return s


def build_filename(fields: List[Optional[str]]) -> str:
    """按指定顺序拼接并返回目标文件名（固定扩展名为.jpg）。"""
    cleaned = [sanitize(f) for f in fields]
    name = "-".join(cleaned) + ".jpg"
    # 防止出现以破折号结尾或开头的异常情况
    name = re.sub(r"^-+", "", name)
    name = re.sub(r"-+\.jpg$", ".jpg", name)
    return name


def strip_query(url: str) -> str:
    """去除URL中的查询参数（?后面的内容）。"""
    if not url:
        return url
    return url.split("?")[0]


def ensure_dir(path: str) -> None:
    """确保输出目录存在。"""
    os.makedirs(path, exist_ok=True)


def find_header_indices(header_row: List[Optional[str]]) -> Dict[str, int]:
    """根据表头名称返回所需列的索引映射。"""
    header_map: Dict[str, int] = {}
    for idx, name in enumerate(header_row):
        if name is None:
            continue
        header_map[str(name).strip()] = idx
    required = ["一级", "二级1", "二级2", "三级", "品牌", "条码", "imageUrl"]
    missing = [col for col in required if col not in header_map]
    if missing:
        raise ValueError(f"Excel表头缺少必要列: {missing}")
    return header_map


def download_file(url: str, dest_path: str, timeout: int = 20) -> bool:
    """下载单个文件，成功返回True，失败返回False。"""
    try:
        resp = requests.get(url, timeout=timeout, stream=True)
        if resp.status_code != 200:
            return False
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except Exception:
        return False


def process_excel(input_path: str, sheet_name: str, output_dir: str, start_row: int = 2, end_row: Optional[int] = None, limit: Optional[int] = None, on_progress: Optional[callable] = None, cancel_event: Optional[object] = None, concurrency: int = 4) -> int:
    """读取Excel并批量下载图片到指定目录。支持进度回调与取消。返回处理记录数。"""
    wb = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}")
    ws = wb[sheet_name]

    # 读取表头
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_idx = find_header_indices(list(header_cells))

    ensure_dir(output_dir)

    # 收集需要处理的记录
    items = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        image_cell = row[header_idx["imageUrl"]]
        if image_cell in (None, ""):
            continue
        one = row[header_idx["一级"]]
        two1 = row[header_idx["二级1"]]
        two2 = row[header_idx["二级2"]]
        three = row[header_idx["三级"]]
        brand = row[header_idx["品牌"]]
        barcode = row[header_idx["条码"]]
        image_url = strip_query(str(image_cell))
        filename = build_filename([one, two1, two2, three, brand, barcode])
        dest_path = os.path.join(output_dir, filename)
        items.append((image_url, dest_path, filename))

    if limit is not None:
        items = items[:limit]

    total = len(items)
    processed = 0
    lock = threading.Lock()

    def handle_skip(filename: str):
        nonlocal processed
        with lock:
            processed += 1
            if on_progress:
                on_progress({"status": "skip", "processed": processed, "total": total, "filename": filename})
            else:
                print(f"已存在，跳过：{os.path.join(output_dir, filename)}")
            write_log(f"已存在，跳过：{filename}")

    def handle_success(filename: str, dest_path: str):
        nonlocal processed
        with lock:
            processed += 1
            if on_progress:
                on_progress({"status": "success", "processed": processed, "total": total, "filename": filename})
            else:
                print(f"下载成功：{dest_path}")
            write_log(f"下载成功：{filename}")

    def handle_fail(filename: str, url: str):
        if on_progress:
            on_progress({"status": "fail", "processed": processed, "total": total, "filename": filename})
        else:
            print(f"下载失败：{url}")
        write_log(f"下载失败：{filename}")
        write_error(f"下载失败：{filename} | URL：{url}")

    # 并发下载
    with ThreadPoolExecutor(max_workers=max(1, int(concurrency or 1))) as executor:
        futures = []
        for url, dest_path, filename in items:
            if cancel_event is not None and getattr(cancel_event, "is_set", lambda: False)():
                break
            if os.path.exists(dest_path):
                handle_skip(filename)
                continue
            def task(u=url, d=dest_path, f=filename):
                if cancel_event is not None and getattr(cancel_event, "is_set", lambda: False)():
                    return ("cancel", f, u, d)
                ok = download_file(u, d)
                return ("success" if ok else "fail", f, u, d)
            futures.append(executor.submit(task))

        for fut in as_completed(futures):
            status, filename, url, dest_path = fut.result()
            if status == "success":
                handle_success(filename, dest_path)
            elif status == "fail":
                handle_fail(filename, url)

    wb.close()
    if on_progress:
        on_progress({"status": "done", "processed": processed, "total": total})
    else:
        print(f"完成，处理记录数：{processed}")
    write_log(f"完成：{processed}/{total}")
    return processed


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(description="从Excel下载图片并重命名保存")
    parser.add_argument("--input", required=True, help="Excel文件路径")
    parser.add_argument("--sheet", default="Sheet1", help="工作表名称")
    parser.add_argument("--out", default="images", help="输出目录")
    parser.add_argument("--start", type=int, default=2, help="开始行（含），默认2跳过表头")
    parser.add_argument("--end", type=int, default=None, help="结束行（含），默认到最后一行")
    parser.add_argument("--limit", type=int, default=None, help="处理记录上限，仅用于试运行")
    parser.add_argument("--concurrency", type=int, default=4, help="并发下载线程数")
    return parser.parse_args(argv)


def main() -> None:
    args = parse_args()
    process_excel(
        input_path=args.input,
        sheet_name=args.sheet,
        output_dir=args.out,
        start_row=args.start,
        end_row=args.end,
        limit=args.limit,
        concurrency=args.concurrency,
    )


if __name__ == "__main__":
    main()